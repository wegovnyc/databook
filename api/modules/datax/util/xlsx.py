import os

import openpyxl

from . import utils


def read(file_path: str, sheet_name: str, title_row: bool):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    result = (('' if value is None else str(value) for value in row) for row in sheet.values)

    if title_row:
        return utils.dicts_from_matrix(result)

    return list(list(row) for row in result)


def clear(file_path: str, sheet_name: str):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        sheet.delete_rows(1, sheet.max_row)
        workbook.save(file_path)
    except (FileNotFoundError, KeyError):
        pass


def append(file_path: str, sheet_name: str, data: [], title_row: bool, titles):
    if not data and not title_row:
        return

    if os.path.isfile(file_path):
        workbook = openpyxl.load_workbook(file_path)
    else:
        parent = os.path.dirname(file_path)
        if parent and not os.path.isdir(parent):
            os.makedirs(parent)
        workbook = openpyxl.Workbook()
        workbook.active.title = sheet_name

    try:
        sheet = workbook[sheet_name]
    except KeyError:
        sheet = workbook.create_sheet(sheet_name)

    if title_row:
        sheet_titles = [] if sheet.max_row == 1 else list(next(sheet.values))

        if titles is None:
            titles = utils.get_titles(data)
        utils.expand_unique(sheet_titles, titles)

        if titles:
            for col_i in range(0, len(sheet_titles)):
                sheet.cell(1, col_i + 1, sheet_titles[col_i])

        for row in data:
            sheet.append({sheet_titles.index(k) + 1: row[k] for k in titles if k in row})
    else:
        for row in data:
            sheet.append(row)

    workbook.save(file_path)
